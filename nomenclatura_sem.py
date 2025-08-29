import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Шляхи до файлів (Windows)
input_nom_path = r'F:\Олег\Семена\Input_nom_sem.xlsx'
ref_path = r'F:\Олег\СЗР\Справочники\Справочник продуктов.xlsx'
output_path = r'F:\Олег\Семена\Output_nom_sem.xlsx'

# 1) Зчитуємо вхідну номенклатуру та необхідні поля
#    "Номенклатура з одиницею виміру", "Номенклатура", "Одиниця виміру", "Код УКТЗЕД", "Тип УКТЗЕД"
df_nom = pd.read_excel(input_nom_path, sheet_name='номенклатура')

# 2) Зчитуємо довідник продуктів (лише Насіння) та необхідні поля
df_ref_raw = pd.read_excel(ref_path, sheet_name='спрПродукты')
df_ref = df_ref_raw[df_ref_raw['Категорія'] == 'Насіння'][
    ['Загальна назва', 'Внутрішній код', 'Культура', 'Субкатегорія', 'Бренд']
]
# Мапа субкатегорії за внутрішнім кодом (для спеціальної логіки Овочеві/Квіти)
subcat_by_code = df_ref.set_index('Внутрішній код')['Субкатегорія']

# 3) Завантажуємо додатковий словник культур та долучаємо до df_nom
df_codes = pd.read_excel(ref_path, sheet_name='спрКоды')[["Код УКТЗЕД", "Культура"]]
if 'Код УКТЗЕД' in df_nom.columns:
    df_nom['Код УКТЗЕД'] = df_nom['Код УКТЗЕД'].astype(str)
    df_codes['Код УКТЗЕД'] = df_codes['Код УКТЗЕД'].astype(str)
    df_nom = df_nom.merge(df_codes, on='Код УКТЗЕД', how='left')
else:
    df_nom['Культура'] = None

# 4) Фільтрація довідника: назви >=4 символів та виключні коди
norm_len = (
    df_ref['Загальна назва'].astype(str)
    .str.replace(r"[^0-9a-zA-Zа-яА-ЯіїєґІЇЄҐ]+", '', regex=True)
    .str.len()
)
df_ref = df_ref[norm_len >= 4]
codes_to_remove = [
    'NROZ-000196','NNEZ-000001','NOGI-000045','NPOZ-000104','NAAA-000003',
    'NGAZ-000004','NSON-000375','NSAL-000074','NPOZ-000142','NBAK-000005',
    'NGOR-000046','NPOM-000066','NKUC-000047','NKAP-000043','NRED-000016',
    'NSAL-000243','NBUK-000011','NKAB-000008','NCIB-000016','NTSC-000011',
    'NBUC-000135','NOGI-000164','NSAL-000245','NMAL-000005','NCHO-000016','NVIO-000010','NPER-000023','NCIK-000006','NPEL-000017','NBAZ-000001'
]
df_ref = df_ref[~df_ref['Внутрішній код'].isin(codes_to_remove)]

# 5) Автозаміни та нормалізація тексту
replacements = {
    'le': 'ле',

    # англ/укр нормалізації назв протруйників (СПОЧАТКУ лексеми!)
    'maxim xl': 'Максим XL',
    'maxim': 'Максим',
    'apron xl': 'Апрон XL',
    'apron': 'Апрон XL',
    'poncho': 'Пончо',
    'force zea': 'Форс Зеа',
    'forcezea': 'Форс Зеа',
    'force_zea': 'Форс Зеа',
    'форс магна': 'Форс Магна',
    'форсмагна': 'Форс Магна',
    'feuver': 'Февер',
    'елівіо енерджі': 'Епівіо Енерджі',
    'елівіо': 'Епівіо',
    'lumigen': 'LumiGEN',
    'луміген': 'LumiGEN',
    'круізер': 'Круїзер',
    'круіз': 'Круїзер',

    # брендові/загальні
    'dkc': 'дкс', 'mas': 'мас', 'квс ': '', 'ес ': '', 'лг ': '', ' f1': '', ' сл':'',
    'redigom': 'Редіго М', 'b360': 'В360', 'stine': 'стайн', 'ржт ': '',' клп':'',' сх':'','си':'','сс':'с','лл':'л','рр':'р',

    # загальні лат→укр літерні заміни (ПІСЛЯ лексем!)
    'й':'ї','и':'і','p':'п','h':'г','e':'е','f':'ф','l':'л',
    'c':'с','i':'і','r':'р','y':'у','d':'д','k':'к','n':'н','s':'с',
    'm':'м','a':'а','u':'у','b':'б','t':'т','x':'х','o':'о','h':'н','v':'в'
}

def normalize_and_replace(text):
    s = str(text).lower()
    # базові чистки
    s = re.sub(r'[-–—]', ' ', s)
    s = s.replace('_', ' ')
    # лапки → пробіли, щоб не склеювало слова ("Рогізнянка"еліта → Рогізнянка еліта)
    s = s.replace('"',' ').replace("'",' ').replace('«',' ').replace('»',' ').replace('“',' ').replace('”',' ')
    # розклеювання типових зліплень перед словниковими замінами
    s = re.sub(r'(?<=\w)(xl|хл)\b', r' \1', s)    # ...xl → ... xl
    s = re.sub(r'maxim\s*xl', 'maxim xl', s)
    s = re.sub(r'force\s*zea', 'force zea', s)    # forcezea/force_zea → force zea
    s = re.sub(r'форс\s*магна', 'форс магна', s)  # форсмагна → форс магна
    s = re.sub(r'круїзер\s*osr', 'круїзер osr', s)

    # СПОЧАТКУ лексеми, потім літерні заміни — порядок ВАЖЛИВИЙ
    for wrong, correct in replacements.items():
        s = s.replace(wrong, correct)

    # фінальна чистка символів
    return re.sub(r'[^0-9a-zа-яіїєґ\+\*/\/,\. _]', '', s)


# 6) Підготовка довідника (нормалізовані назви)
df_ref['norm_name'] = df_ref['Загальна назва'].apply(normalize_and_replace)
df_ref = df_ref.sort_values('norm_name', key=lambda c: c.str.len(), ascending=False)

# 7) Карта одиниць обчислення та коефіцієнти для культур
unit_map = {
    'т': 1, 'т.': 1, 'кг': 1, 'кгмістк': 1, 'ц': 1, 'г': 1
}
special_coeff = {
    'Кукурудза': 37.3134, 'Соняшник': 95.2381, 'Буряк цукровий': 571.4286,
    'Жито': 28.5714, 'Ріпак озимий': 126.9841, 'Ріпак ярий': 126.9841,
    'Сорго': 114.9425
}

# Культури, для яких у разі "нестандартної" одиниці зчитуємо масу з тексту номенклатури
MASS_CULTURES = {
    'Горох','Горох озимий','Гречка','Льон','Люпин','Нут',
    'Овес','Просо','Пшениця озима','Пшениця яра','Сочевиця',
    'Соя','Тритікале','Ячмінь озимий','Ячмінь ярий'
}

def _extract_mass_from_text(nom_text: str):
    """Повертає масу у тоннах, якщо у тексті є число + одиниця (т/кг/ц/г).
    Підтримує десяткову кому або крапку, допускає пробіли та дужки.
    Приклад: "(0,025 т)" → 0.025; "0.025 кг" → 0.000025
    """
    s = normalize_and_replace(nom_text)
    # прибираємо дужки, щоб токени не злипались
    s = s.replace('(', ' ').replace(')', ' ')
    parts = s.split()
    for i, tok in enumerate(parts):
        t = tok.replace(',', '.').strip()
        # злиплі варіанти типу 0,025т / 0.025кг / 0.025ц / 0.025г
        if t.endswith('т') or t.endswith('т.'):
            num = t[:-1].rstrip('.')
            try:
                return float(num)
            except:
                pass
        if t.endswith('кг'):
            num = t[:-2]
            try:
                return float(num) / 1000.0
            except:
                pass
        if t.endswith('ц'):
            num = t[:-1]
            try:
                return float(num) / 10.0
            except:
                pass
        if t.endswith('г'):
            num = t[:-1]
            try:
                return float(num) / 1_000_000.0
            except:
                pass
        # варіант: число та одиниця окремими токенами
        try:
            val = float(t)
        except:
            val = None
        if val is not None and i + 1 < len(parts):
            u = parts[i+1]
            if u.startswith('т'):
                return val
            if u == 'кг':
                return val / 1000.0
            if u == 'ц':
                return val / 10.0
            if u == 'г':
                return val / 1_000_000.0
    return None

# 8) Пошук внутрішнього коду
def match_code(nom_text, culture, uktzed):
    desc = normalize_and_replace(nom_text)

    # Heuristic: виправлення кодів типу Р64LE185 → П64ЛЕ185 для пошуку в довіднику
    def _fix_pioneer_token(tok: str) -> str:
        t = tok.replace('-', '').replace('_', '').replace(' ', '')
        if len(t) >= 5:
            first = t[0]
            if first in ['p', 'р'] and t[1].isdigit() and t[2].isdigit():
                rest = t[3:]
                if rest.startswith('le'):
                    rest = 'ле' + rest[2:]
                if rest.startswith('ле'):
                    return 'п' + t[1] + t[2] + rest
        return tok

    toks = desc.split()
    toks = [_fix_pioneer_token(x) for x in toks]
    desc = ' '.join(toks)

    # Синонім: помідор → томат (у різних формах)
    desc_pad = ' ' + desc + ' '
    for a in ['помідор', 'помідори', 'помидор', 'помидоры']:
        desc_pad = desc_pad.replace(f' {a} ', ' томат ')
    desc = ' '.join(desc_pad.split())

    search = desc.replace(',', '').replace('.', '').replace(' ', '')

    sub = df_ref.copy()
    veg_flower_uktzeds = ['1209918000','1209919000','1209300000','1209']
    is_veg = uktzed in ['1209918000','1209919000']
    is_flower = uktzed == '1209300000'
    is_veg_or_flower = uktzed in veg_flower_uktzeds

    if is_veg:
        sub = sub[sub['Субкатегорія'] == 'Овочеві']
    elif is_flower:
        sub = sub[sub['Субкатегорія'] == 'Квіти']
    elif uktzed in ['1003','1001110000','1003100000']:
        sub = sub[sub['Культура'].isin(['Ячмінь озимий','Ячмінь ярий'])]
    elif uktzed in ['1001','1001912000']:
        sub = sub[sub['Культура'].isin(['Пшениця озима','Пшениця яра'])]
    elif uktzed in ['1207700000']:
        sub = sub[sub['Культура'].isin(['Диня','Кавун'])]
    elif uktzed in ['713101000']:
        sub = sub[sub['Культура'].isin(['Горох','Горох овочевий'])]
    elif uktzed == '1209':
        sub = sub[sub['Субкатегорія'].isin(['Квіти','Овочеві'])]
    else:
        sub = sub[sub['Культура'] == culture]

    # ДОДАТКОВО для Овочевих/Квітів: спробувати виявити культуру за назвою у номенклатурі
    if is_veg_or_flower:
        # Унікальні культури в межах поточного підмноження
        cand_cultures = [c for c in sub['Культура'].dropna().unique()]
        found_cult = None
        desc_padded = ' ' + desc + ' '
        for c in cand_cultures:
            norm_c = normalize_and_replace(c)
            if f' {norm_c} ' in desc_padded:
                found_cult = c
                break
        if found_cult:
            sub = sub[sub['Культура'] == found_cult]

    for _, row in sub.iterrows():
        norm = row['norm_name']
        if (norm in desc) or (norm.replace(' ', '') in search):
            return row['Внутрішній код']
    return None

# 9) Новий розрахунок коефіцієнта базової одиниці
def calc_coeff(nom_text, unit_text, culture, inner_code):
    u = str(unit_text).lower().replace(' ', '').replace('.', '')
    base = special_coeff.get(culture, 1)
    if u in ['т', 'ттовпрод']:
        return base
    if u.startswith('кг'):
        specific = ['Горох','Горох озимий','Гречка','Льон','Люпин','Нут',
                    'Овес','Просо','Пшениця озима','Пшениця яра','Сочевиця',
                    'Соя','Тритікале','Ячмінь озимий','Ячмінь ярий']
        if culture in specific: return 0.001
        if culture not in special_coeff: return 1
        return base / 1000
    if u == 'ц':
        specific = ['Горох','Горох озимий','Гречка','Льон','Люпин','Нут',
                    'Овес','Просо','Пшениця озима','Пшениця яра','Сочевиця',
                    'Соя','Тритікале','Ячмінь озимий','Ячмінь ярий']
        if culture in specific: return 0.1
        return base / 10
    if u == 'г':
        specific = ['Горох','Горох озимий','Гречка','Льон','Люпин','Нут',
                    'Овес','Просо','Пшениця озима','Пшениця яра','Сочевиця',
                    'Соя','Тритікале','Ячмінь озимий','Ячмінь ярий']
        if culture in specific: return 0.000001
        return base / 1e6

    # Якщо одиниця виміру НЕ з [т/кг/ц/г], але культура з MASS_CULTURES — пробуємо витягнути масу з тексту
    if culture in MASS_CULTURES and not (u in ['т','ттовпрод'] or u.startswith('кг') or u in ['ц','г']):
        parsed = _extract_mass_from_text(nom_text)
        if parsed is not None:
            return parsed
    # Овочеві/Квіти: як MASS_CULTURES, але у 1000 разів більше (кг-база)
    subcat = subcat_by_code.get(inner_code) if 'subcat_by_code' in globals() else None
    if subcat in ['Овочеві','Квіти'] and not (u in ['т','ттовпрод'] or u.startswith('кг') or u in ['ц','г']):
        parsed_vk = _extract_mass_from_text(nom_text)
        if parsed_vk is not None:
            return parsed_vk * 1000.0

    txt = normalize_and_replace(nom_text)

    if culture == 'Кукурудза':
        if inner_code in df_ref_raw['Внутрішній код'].values:
            brand = df_ref_raw.loc[df_ref_raw['Внутрішній код'] == inner_code, 'Бренд'].iloc[0]
            if brand in ['KWS','LG Seeds','MAS Seeds','RAGT','Golden West Seeds']:
                return 0.625
        if '25000' in txt or '25тис' in txt or '25мк' in txt: return 0.3125
        if '70тис' in txt: return 0.875
        if 'д50' in txt: return 0.625
        return 1

    if culture == 'Соняшник':
        if 'd75' in txt: return 0.5
        if '25000' in txt: return 0.1667
        return 1

    if culture == 'Жито':
        if re.search(r'\b25\s*м[нn]\b', txt):
            return 25
        return 1

    if culture in ['Ріпак озимий','Ріпак ярий']:
        # 750 тис. (із пробілом або без)
        if '750тис' in txt or '750 тис' in txt: return 0.5
        # Якщо присутні коди PX/ПХ → Pioneer
        if ' px' in (' ' + txt) or ' пх' in (' ' + txt):
            return 1.3333
        matches = df_ref[df_ref['norm_name'] == txt]
        if not matches.empty and matches['Бренд'].iloc[0] == 'Pioneer': return 1.3333
        return 1

    if culture == 'Сорго':
        if '20кг' in txt: return 2.2989
        return 1

    if culture in ['Пшениця озима','Пшениця яра']:
        if '50кг' in txt: return 0.05
        if '500кг' in txt: return 0.5
        return 1

    if culture in ['Ячмінь озимий','Ячмінь ярий']:
        if '750кг' in txt: return 0.75
        if '10млн' in txt or '10м ' in txt: return 0.45
        if '500к' in txt: return 0.0235
        return 1

    if culture == 'Соя':
        return 0.025

    return 1
    if culture == 'Соняшник':
        if 'd75' in txt: return 0.5
        if '25000' in txt: return 0.1667
        return 1
    if culture == 'Жито':
        if '25мн' in txt: return 25
        return 1
    if culture in ['Ріпак озимий','Ріпак ярий']:
        if '750тис' in txt: return 0.5
        matches = df_ref[df_ref['norm_name'] == txt]
        if not matches.empty and matches['Бренд'].iloc[0] == 'Pioneer': return 1.3333
        return 1
    if culture == 'Сорго':
        if '20кг' in txt: return 2.2989
        return 1
    if culture in ['Пшениця озима','Пшениця яра']:
        if '50кг' in txt: return 0.05
        if '500кг' in txt: return 0.5
        return 1
    if culture in ['Ячмінь озимий','Ячмінь ярий']:
        if re.search(r'10млн|10м\b', txt): return 0.45
        if '500к' in txt: return 0.0235
        return 1
    if culture == 'Соя': return 0.025
    return 1

# 10) Застосування обчислень до DataFrame
# Обчислюємо внутрішній код
df_nom['Внутрішній код'] = df_nom.apply(
    lambda r: match_code(r['Номенклатура'], r.get('Культура',''), r.get('Код УКТЗЕД','')),
    axis=1
)
# Обчислюємо коефіцієнт базової одиниці
df_nom['Коефіцієнт базової одиниці'] = df_nom.apply(
    lambda r: calc_coeff(r['Номенклатура'], r.get('Одиниця виміру',''), r.get('Культура',''), r.get('Внутрішній код','')),
    axis=1
)

# Обмеження застосування логіки за культурами
PROTECTANT_CULTURES = {
    'Соняшник','Кукурудза','Буряк цукровий','Ріпак озимий','Пшениця озима',
    'Жито','Соя','Ячмінь озимий','Кукурудза цукрова','Ячмінь ярий',
    'Пшениця яра','Сорго','Ріпак ярий','Горох','Газонні трави'
}
GENERATION_CULTURES = {
    'Пшениця озима','Соя','Ячмінь озимий','Ячмінь ярий','Пшениця яра','Горох',
    'Ріпак озимий','Овес','Гречка','Люцерна','Соняшник','Горох озимий',
    'Пшениця дворучка','Просо','Жито','Тритікале','Льон','Нут','Кукурудза','Ріпак ярий'
}

# 11) Препарати-протруйники
pro_list_orig = [
    'Круїзер OSR','Круїзер','Форс Зеа','Вайбранс Інтеграл','Вайбранс',
    'Епівіо Енерджі','Акселерон Еліт','Акселерон Стандарт',
    'LumiGEN','Пончо','Пончо Бета','Модесто Плюс','Максим XL','В360',
    'Апрон XL','Редіго М','Boost&Go','Кінто Дуо','Кінто Плюс','Гаучо Плюс',
    'Агростарт','Торк СТ','Форс Магна','PB1','Вітазим','Рексолін АВС',
    'Гімексазол','Тіаметоксам','Тефлутрін','Радікс Перфекта','Аліос',
    'Лумісена','Вітавакс','АскоСтарт','Сімбіоз','Біостимулятор','ДуоМікс',
    'Інітіо Про','Тірам','Флудіоксоніл','Металаксил','Седаксан','Сульфо',
    'Февер','Селест Макс','Клотіанідин','Форс','Апрон','Максим','Іншур Перформ','Командор Гранд','Пончо Вотіво','Оплот Тріо','Табу'
]

pro_list_orig.append('Роялфло')
# Канонічна нормалізація суто для пошуку протруйників (без regex у викликах заміни)
def canon_protectant_text(text):
    s = str(text).lower()
    # базові заміни символів розділювачів
    for ch in ['_', '-', '–', '—']:
        s = s.replace(ch, ' ')
    # прибираємо зайву пунктуацію, щоб не заважала словним межам
    for ch in [',', '.', '+', '(', ')', '[', ']', '{', '}', ':', ';', '"', "'", '!', '?', '*']:
        s = s.replace(ch, ' ')
    # додатково: & → пробіл, склейки boostgo → boost go
    s = s.replace('&', ' ')
    s = s.replace('boostgo', 'boost go')
    # нормалізація ключових лексем і синонімів
    s = ' '.join(s.split())
    s = s.replace('forcezea', 'force zea').replace('force zea', 'форс зеа')
    s = s.replace('force zea280', 'форс зеа')
    s = s.replace('форс магна', 'форс магна').replace('форсмагна', 'форс магна')
    s = s.replace('feuver', 'февер')
    s = s.replace('elivio', 'епівіо').replace('epivio', 'епівіо').replace('елівіо', 'епівіо')
    s = s.replace('lmg', 'lumigen').replace('луміген', 'lumigen')
    s = s.replace('cruiser', 'круїзер').replace('круізер', 'круїзер')
    s = s.replace('agrostart', 'агростарт')
    s = s.replace('royal flo', 'роялфло').replace('роял фло', 'роялфло')
    # вирівнюємо maxim/apron/poncho (враховуємо латинсько-кириличну "і")
    s = s.replace('мaxim', 'maxim').replace('maxім', 'maxim')
    s = s.replace('maximxl', 'maxim xl').replace('maxim xl', 'максим xl').replace('maxim xl', 'максим xl')
    s = s.replace('apronxl', 'apron xl').replace('apron xl', 'апрон xl').replace('apron', 'апрон xl')
    s = s.replace('poncho', 'пончо')
    # обробка випадку "maxim <число>" → "максим"
    tokens = s.split()
    out = []
    i = 0
    while i < len(tokens):
        tok = tokens[i]
        if tok == 'maxim':
            nxt = tokens[i+1] if i+1 < len(tokens) else ''
            if nxt.replace('.', '').isdigit():
                out.append('максим')
                i += 2
                continue
            else:
                out.append('максим')
                i += 1
                continue
        out.append(tok)
        i += 1
    s = ' '.join(out)
    # фінальне ущільнення пробілів
    s = ' '.join(s.split())
    return s

# Створюємо список (канонічна_форма, оригінальна_назва) і сортуємо за довжиною
pro_list = [(canon_protectant_text(p), p) for p in pro_list_orig]
pro_list.sort(key=lambda x: len(x[0]), reverse=True)

# Пошук: довші збіги мають пріоритет, використовуємо словні межі через пробіли
def find_protectants(nom):
    # Попередня нормалізація для кращого знаходження склеєних форм і абревіатур БЕЗ regex у замінах
    pre = str(nom).lower()
    for ch in ['+', '_', '-', '–', '—', ',', '.', '(', ')', '[', ']', '{', '}', ':', ';', '"', "'", '!', '?', '*']:
        pre = pre.replace(ch, ' ')
    pre = pre.replace('&', ' ')
    pre = ' '.join(pre.split())

    # розклеїти склеєні конструкції та синоніми
    pre = pre.replace('boostgo', 'boost go')
    pre = pre.replace('максимxl', 'максим xl').replace('апронxl', 'апрон xl')
    for d in '0123456789':
        pre = pre.replace('xl'+d, 'xl '+d)
        pre = pre.replace('плюс'+d, 'плюс '+d)
    # абревіатури/синоніми брендів
    pre = ' ' + pre + ' '
    pre = pre.replace(' фз ', ' форс зеа ')
    pre = pre.replace(' фм ', ' форс магна ')
    pre = pre.replace(' acceleron ', ' акселерон ')
    pre = pre.replace(' standart ', ' стандарт ')
    pre = pre.replace(' standard ', ' стандарт ')
    pre = pre.replace(' лумінг ', ' lumigen ')
    pre = pre.replace(' luming ', ' lumigen ')
    pre = pre.replace(' lmg ', ' lumigen ')
    pre = ' '.join(pre.split())

    # --- Пріоритет для Акселерон: нормалізуємо "Elite" → "Еліт"
    pre = ' ' + pre + ' '
    pre = pre.replace(' elite ', ' еліт ').replace(' elit ', ' еліт ')
    pre = ' '.join(pre.split())

    txt = ' ' + canon_protectant_text(pre) + ' '

    # --- Якщо є Акселерон Еліт/Стандарт — повертаємо ТІЛЬКИ їх
    acc_only = []
    if ' акселерон еліт ' in txt:
        acc_only.append('Акселерон Еліт')
    if ' акселерон стандарт ' in txt:
        acc_only.append('Акселерон Стандарт')
    if acc_only:
        return '/'.join(acc_only)

    found = []
    temp = txt
    for canon, orig in pro_list:
        if not canon:
            continue
        segment = ' ' + canon + ' '
        if segment in temp:
            found.append(orig)
            temp = temp.replace(segment, ' ')
    # додаткова морфологія для "круїзер(ом/у/а/и/і/ем)"
    words = temp.split()
    if any(w.startswith('круїзер') for w in words):
        if 'Круїзер' not in found:
            found.append('Круїзер')

    # унікалізація з порядком
    unique = []
    for p in found:
        if p not in unique:
            unique.append(p)
    if unique:
        return '/'.join(unique)
    base_txt = txt.strip()
    if any(k in base_txt for k in ['непротр','н/прот','не протр']):
        return ''
    if any(k in base_txt for k in ['прот','інс','фун']):
        return 'Протруєне (НД)'
    return ''
    if any(k in base_txt for k in ['прот','інс','фун']):
        return 'Протруєне (НД)'
    return ''
    if any(k in base_txt for k in ['прот','інс','фун']):
        return 'Протруєне (НД)'
    return ''

df_nom['Препарати-протруйники'] = df_nom.apply(
    lambda r: find_protectants(r['Номенклатура']) if str(r.get('Культура','')) in PROTECTANT_CULTURES else '',
    axis=1
)

# 12) Демо-статус
# Якщо в назві є 'демо' або 'demo' -> 'Демо'; якщо одиниця виміру у переліку вагових одиниць -> 'Вагове'; інакше 'Комерційне'
def determine_demo_status(nom, unit):
    txt = str(nom).lower()
    if 'демо' in txt or 'demo' in txt:
        return 'Демо'
    u = str(unit).lower().replace(' ', '')
    if u in ['т','кг','ц','г','ттов.прод','т.','кгмістк']:
        return 'Вагове'
    return 'Комерційне'
df_nom['Демо-статус'] = df_nom.apply(lambda r: determine_demo_status(r['Номенклатура'], r.get('Одиниця виміру','')), axis=1)

# 13) Генерація — ОНОВЛЕНО з урахуванням кейсів:
#  - "І-репродукція" → р1 (римські/українські римські цифри)
#  - "1-ша репродукція" → р1 (число + словесний суфікс)
#  - Игноруємо англ. "Elite" (напр., "Acceleron Elite") при визначенні "Еліта"

def determine_generation(nom):
    raw = str(nom)
    raw_l = raw.lower()

    # 0) Якщо є англ. слово "elite" — НЕ вважати це "Еліта"
    has_english_elite = bool(re.search(r'\belite\b', raw_l))

    txt = normalize_and_replace(raw)

    # 1) Супереліта/Еліта (допускаємо пробіли/дефіси/лапки між словами)
    if re.search(r'\bсупер\W*еліта\b', txt):
        return 'Супереліта'
    if (not has_english_elite) and re.search(r'\bеліта\w*\b', txt):  # еліта/елітна, але не "Elite"
        return 'Еліта'

    # 2) "сн-1", "сн 2", "сн3" → р1..р5
    m = re.search(r'\bсн\W*([1-5])\b', txt)
    if m:
        return f'р{m.group(1)}'

    # 3) Римські цифри перед "репр": "І-репродукція", "ІІ репр", "III-репр."
    #    Враховуємо латинські I,V та українські І (великі/малі)
    m = re.search(r'\b([iv]{1,3}|iv|v|[і]{1,3})\b\s*[-–—]?\s*ре+п*р', raw_l)
    if m:
        token = m.group(1)
        token = token.replace('і', 'i')  # уніфікуємо українську "і" до латинської
        roman_map = {'i':1,'ii':2,'iii':3,'iv':4,'v':5}
        if token in roman_map:
            return f"р{roman_map[token]}"

    # 4) Число з можливим суфіксом "-ша" перед словом "репр": "1-ша репродукція", "2ша репр."
    m = re.search(r'\b([1-5])\s*[-–—]?\s*(?:ша)?\s*репр', raw_l)
    if m:
        return f'р{m.group(1)}'

    # 5) "1 репродукція" / "2репр" / "3 репрод." → р1..р5 (пробіл необов’язковий)
    m = re.search(r'\b([1-5])\s*ре+п*р\w*', txt)
    if m:
        return f'р{m.group(1)}'

    # 6) Словесні порядкові: "перша/друга/третя/четверта/п’ята репродукція"
    word_map = {'перша':'1','друга':'2','третя':'3','четверта':'4','п’ята':'5',"п'ята":'5','пята':'5'}
    for w, n in word_map.items():
        if re.search(rf'\b{w}\b.*\bрепр', txt):
            return f'р{n}'

    # 7) "G1".."G5" → р1..р5 (англ. позначення генерації)
    m = re.search(r'\bg\W*([1-5])\b', txt)
    if m:
        return f'р{m.group(1)}'

    return ''


def _determine_generation_wrap(nom):
    # спочатку базова логіка
    g = determine_generation(nom)
    if g:
        return g
    # додаткові евристики без regex
    s = str(nom).lower()
    for ch in [',', '.', '+', '(', ')', '[', ']', '{', '}', ':', ';', '"', "'", '!', '?', '*', '-', '–', '—', '/']:
        s = s.replace(ch, ' ')
    s = ' '.join(s.split())
    toks = s.split()

    # швидка перевірка: 1 реп(р/пр...) та 1репр* (без пробілу)
    for idx, t in enumerate(toks):
        if t in ['1','2','3','4','5'] and idx+1 < len(toks):
            nxt = toks[idx+1]
            nxt_fix = nxt
            while 'пп' in nxt_fix:
                nxt_fix = nxt_fix.replace('пп','п')
            if nxt_fix.startswith('репр'):
                return f"р{t}"
    for n in ['1','2','3','4','5']:
        if any(tok.startswith(n + 'реп') or tok.startswith(n + 'репп') for tok in toks):
            return f'р{n}'

    # с/еліта, с/ел → Супереліта
    if 'с/еліта' in s or 'с/е' in s:
        return 'Супереліта'

    # 1р..5р або r1..r5 (як окремі токени)
    for i in ['1','2','3','4','5']:
        if (i + 'р') in toks or ('r' + i) in toks:
            return f'р{i}'

    # СН перша/друга/... (словесні порядкові після СН)
    ord_map = {'перша':'1','друга':'2','третя':'3','четверта':'4','п’ята':'5',"п'ята":'5','пята':'5'}
    for idx, t in enumerate(toks):
        if t == 'сн' and idx+1 < len(toks) and toks[idx+1] in ord_map:
            return f"р{ord_map[toks[idx+1]]}"

    # СН1..СН5 (без пробілу)
    for t in toks:
        if t in ['сн1','сн2','сн3','сн4','сн5']:
            return f"р{t[-1]}"

    # 1генерац / 2генерац ... (без пробілу) або "1 генерац*" (з пробілом)
    for idx, t in enumerate(toks):
        if t in ['1','2','3','4','5'] and idx+1 < len(toks) and toks[idx+1].startswith('генерац'):
            return f"р{t}"
    for n in ['1','2','3','4','5']:
        if any(tok.startswith(n + 'генерац') for tok in toks):
            return f'р{n}'

    # Р-2 / РН-2 (кирилична р як окремий токен перед числом)
    for idx, t in enumerate(toks):
        if t in ['р', 'рн'] and idx+1 < len(toks) and toks[idx+1] in ['1','2','3','4','5']:
            return f"р{toks[idx+1]}"

    # І р. → р1 (також ii/iii перед 'р')
    romans = {'і':1,'ii':2,'іі':2,'iii':3,'ііі':3}
    for idx, t in enumerate(toks[:-1]):
        if t in romans and toks[idx+1] == 'р':
            return f"р{romans[t]}"

    return ''

df_nom['Генерація'] = df_nom.apply(
    lambda r: _determine_generation_wrap(r['Номенклатура']) if str(r.get('Культура','')) in GENERATION_CULTURES else '',
    axis=1
)

# 12) Запис результату та підсвічування нових колонок
# Видаляємо колонку 'Культура' з результату
if 'Культура' in df_nom.columns:
    df_nom.drop(columns=['Культура'], inplace=True)

# Безпечний запис результату з fallback, якщо файл відкрито в Excel
save_path = output_path
try:
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        df_nom.to_excel(writer, sheet_name='номенклатура', index=False)
except PermissionError:
    from datetime import datetime
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    save_path = output_path.replace('.xlsx', f'_{ts}.xlsx')
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        df_nom.to_excel(writer, sheet_name='номенклатура', index=False)

wb = load_workbook(save_path)
ws = wb['номенклатура']
from openpyxl.styles import Font
# Заголовки: робимо жирним шрифт для всіх колонок
for cell in ws[1]:
    cell.font = Font(bold=True)
# Текст у нових колонках: колір #5D489A
highlight_font = Font(color='5D489A')
input_cols = ['Номенклатура з одиницею виміру', 'Номенклатура', 'Одиниця виміру', 'Код УКТЗЕД', 'Тип УКТЗЕД']
for idx, cell in enumerate(ws[1], 1):
    if cell.value not in input_cols:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=idx).font = highlight_font
wb.save(save_path)

print('Готово! Результат збережено в', output_path)
