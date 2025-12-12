import pandas as pd
import re
import numpy as np

# ЖОРСТКИЙ патерн для запчастин
pattern_strict = r"""(?iu)
\b(?:
    втулк\w*|
    пал[еє]ц\w*|
    в[іi]сь\w*|
    вал(?!ют)\w*|
    гайк\w*|болт\w*|шайб\w*|
    к[іi]льц\w*|манжет\w*|прокладк\w*|упор\w*|
    муфт\w*|з'єднан\w*|зчеплен\w*|
    корпус(?!\s*по[iі]л)|
    підшипник\w*|вузл(?!\s*охорон)|опор(?!\s*систем)|маточин\w*|
    диск(?!\s*апаратів)|шк[іi]в\w*|з[іi]рочк\w*|колес(?!\s*тягач)|сегмент\w*|решет\w*|сит\w*|
    кронштейн\w*|ст[іi]йк\w*|стойк\w*|башмак\w*|балк(?!\s*транспорту)|рамк(?!\s*докум)|
    плит\w*|борт(?!\s*тягач)|боковин\w*|
    притискувач\w*|натягувач\w*|фіксатор\w*|утримувач\w*|прижим\w*|пильовик\w*|щитк\w*|
    редуктор\w*|каток\w*|апарат(?!\s*захист)|дозатор\w*|розпилювач\w*|патрубок\w*|тройник\w*
)\b
|
\b(?:в|у)\s*збор[іi]\b|
\bсб\.?\b
"""


# Шляхи до файлів (Windows)
input_nom_path = r'F:\Олег\Сельхозтехника\Input_nom_SGT.xlsx'
ref_path = r'F:\Олег\СЗР\Справочники\Справочник продуктов.xlsx'
output_path = r'F:\Олег\Сельхозтехника\Output_nom_SGT.xlsx'

# ------------------------ допоміжні функції ------------------------
def clean_code(x):
    """Нормалізує код УКТЗЕД/УКТЗЕД до рядка з цифр без пробілів та розділювачів."""
    s = str(x) if pd.notna(x) else ''
    s = re.sub(r'[^0-9]', '', s)
    return s

replacements = {'мега':'mega','lexion':'лексіон','мтз':'беларус','diamant':'діамант','rubin':'рубін',
                'dominator':'домінатор','gaspardo':'maschio gaspardo',
                'й':'ї','и':'і','t':'т','r':'р','v':'в','o':'о','a':'а','c':'с','p':'р','m':'м','f':'ф','e':'е'}

def normalize_and_replace(text):
    """
    Нормалізація для пошуку:
    - lower;
    - автозаміни;
    - прибираємо зайві символи;
    - ВАЖЛИВО: видаляємо УСІ пробіли та коми/крапки (щоб 'Maestro 24SW' ~ 'Maestro 24 SW').
    """
    s = str(text).lower()
    s = re.sub(r'[-–—]', ' ', s)
    for wrong, correct in replacements.items():
        s = s.replace(wrong, correct)
    s = re.sub(r'[^0-9a-zа-яіїєґ\+\*/\/,\.× ]', '', s)
    s = re.sub(r'\s+', '', s)
    s = s.replace(',', '').replace('.', '')
    return s

def strip_brand_prefix(name, brand):
    """Вирізає бренд на початку 'Загальна назва' (нечутливо до регістру)."""
    if pd.isna(name) or pd.isna(brand):
        return name
    pattern = r'^\s*' + re.escape(str(brand)).strip() + r'[\s\-:]*'
    return re.sub(pattern, '', str(name), flags=re.IGNORECASE)

def is_digits_only_name(s: str) -> bool:
    """True, якщо в рядку немає літер, але є принаймні одна цифра (допускаємо пробіли/розділювачі)."""
    if pd.isna(s):
        return False
    txt = str(s)
    # якщо є будь-які букви — не підходить
    if re.search(r'[a-zA-Zа-яА-Яіїєґ]', txt):
        return False
    # має бути хоч одна цифра
    return bool(re.search(r'\d', txt))

def extract_digit_tokens(s: str):
    """Витягує числові токени: 12, 108, 24.5, 770, тощо."""
    if pd.isna(s):
        return []
    return re.findall(r'\d+(?:[.,]\d+)?', str(s))

def digits_only_sequence_matches(nom_text: str, tokens) -> bool:
    """
    Для 'числових' назв: підтверджуємо збіг, якщо:
    - кожен сусідній токен у 'Загальній назві' в Номенклатурі розділений НЕцифровим символом;
    - по краях збігу немає прилягання до інших цифр.
    Приклади, що МАЮТЬ збіг: '24 36', '24x36', '24-SW-36', '400'
    Приклади, що НЕ мають збігу: '2436' (бо токени прилягають цифрами), '1400'/'4000' (бо прилягання цифр з краю).
    """
    if not tokens:
        return False
    pattern = r'(?<!\d)' + r'\D+'.join(re.escape(t) for t in tokens) + r'(?!\d)'
    return re.search(pattern, str(nom_text), flags=re.IGNORECASE) is not None


# ------------------------ 1) Вхідні дані ------------------------
df_nom = pd.read_excel(input_nom_path, sheet_name='номенклатура')

# Визначаємо назву колонки з кодом УКТЗЕД у вхідному файлі
code_col_candidates = ['Код УКТЗЕД', 'Тип УКТЗЕД']
code_col = next((c for c in code_col_candidates if c in df_nom.columns), None)
if code_col is None:
    raise KeyError("У файлі Input очікується колонка 'Код УКТЗЕД' або 'Тип УКТЗЕД'.")

df_nom['_code_norm'] = df_nom[code_col].apply(clean_code)

# ------------------------ 2) Довідники ------------------------
# 2.1 «спрКоды»: Код УКТЗЕД → Культура
df_kody = pd.read_excel(ref_path, sheet_name='спрКоды')
if not {'Код УКТЗЕД', 'Культура'}.issubset(df_kody.columns):
    raise KeyError("У листі 'спрКоды' повинні бути колонки 'Код УКТЗЕД' і 'Культура'.")
df_kody['_code_norm'] = df_kody['Код УКТЗЕД'].apply(clean_code)
df_kody = df_kody[['Культура', '_code_norm']].drop_duplicates()

# 2.2 «спрПродукти_СГТ»: довідник продуктів
df_ref_raw = pd.read_excel(ref_path, sheet_name='спрПродукти_СГТ')
required_cols = {'Загальна назва', 'Внутрішній код', 'Субкатегорія', 'Бренд'}
missing = required_cols - set(df_ref_raw.columns)
if missing:
    raise KeyError(f"У листі 'спрПродукти_СГТ' відсутні колонки: {sorted(missing)}")


# --- НОВЕ: фільтр за «небажаними» підрядками у 'Загальна назва'
bad_substrings = [
    "агрегат","аератор","асенізац","борон","бункер","буртоукл","буряко","викорч","візок","газонокос",
    "гичкозб","глибокороз","грабл","грядкоут","грунтофрез","дощувал","екскаватор","навантажувач","жнивар",
    "зворушувач","зерносуш","змішувач","зчіпн","картоплек","каток","комбайн","копалк","кормозм","кормороз",
    "корчувач","косарк","культиватор","переробк","розрізан","міксер","молотарк","мотоблок","мульчувач","дрон",
    "обертач","компост","обмотувал","обприскувач","обрізувач","очисн","підрізчик","плуг","подрібн","підбирач",
    "пересадк","причеп","причіп","протруювач","розкидач","розпушувач","саджалк","зрошен","системонос","сівалк",
    "скарифікатор","сортувал","трактор","ущільнювач","фреза","цистерн","компакт"
]

# НЕзахоплююча група (?:...) + ігнор регістру
pattern_bad = re.compile(r'(?:' + '|'.join(map(re.escape, bad_substrings)) + r')', flags=re.IGNORECASE)

# тепер без попередження
mask_bad = df_ref_raw['Загальна назва'].astype(str).str.contains(pattern_bad, na=False)
removed_count = int(mask_bad.sum())
df_ref_raw = df_ref_raw[~mask_bad].copy()

print(f"Видалено рядків за стоп-словами: {removed_count}")


# --- Вирізання бренду на початку назви (перед аналізом)
df_ref_raw['Назва_для_аналізу'] = df_ref_raw.apply(
    lambda r: strip_brand_prefix(r['Загальна назва'], r['Бренд']),
    axis=1
)

# --- Попереднє сортування за довжиною очищеної назви (від більшої до меншої)
df_ref_raw = df_ref_raw.sort_values(
    'Назва_для_аналізу',
    key=lambda s: s.astype(str).str.len(),
    ascending=False
).reset_index(drop=True)

# ------------------------ 3) Матчимо Культуру до номенклатури ------------------------
df_nom = df_nom.merge(df_kody, on='_code_norm', how='left')  # додаємо 'Культура'

# ------------------------ 4) Підготовка довідника ------------------------
df_ref = df_ref_raw[['Назва_для_аналізу', 'Внутрішній код', 'Субкатегорія']].copy()

norm_len = (df_ref['Назва_для_аналізу'].astype(str)
            .str.replace(r"[^0-9a-zA-Zа-яА-Яіїєґ]+", '', regex=True)
            .str.len())
df_ref = df_ref[norm_len >= 4]

codes_to_remove = ['TBDI-000687','TKZE-001040','TMUL-000055','TKOV-000157','TTKL-000571','TZZE-000254','TSPR-000639',
                   'TPLU-000577','TMOT-000385','TKZE-001235','TVIZ-000184','TMOT-000420','TZKU-000398','TVIZ-000504',
                   'TSPR-000851','TTKL-001784','TKOS-000315','TGLI-000294','TKUL-001494']
df_ref = df_ref[~df_ref['Внутрішній код'].isin(codes_to_remove)].reset_index(drop=True)

# Нормалізована назва (без пробілів, у lower) для substring-пошуку
df_ref['norm_name'] = df_ref['Назва_для_аналізу'].apply(normalize_and_replace)

# Службові ознаки для числових назв
df_ref['is_digits_only'] = df_ref['Назва_для_аналізу'].apply(is_digits_only_name)
df_ref['digit_tokens']   = df_ref['Назва_для_аналізу'].apply(extract_digit_tokens)

# Індекс за субкатегоріями (з нужними колонками)
cat_to_ref = {
    cat: sub_df[['Внутрішній код', 'norm_name', 'is_digits_only', 'digit_tokens']].reset_index(drop=True)
    for cat, sub_df in df_ref.groupby('Субкатегорія', sort=False)
}

# Глобальний список для fallback-пошуку (також з нужними колонками)
df_ref_all = df_ref[['Внутрішній код', 'norm_name', 'Субкатегорія', 'is_digits_only', 'digit_tokens']].reset_index(drop=True)


# ------------------------ 5) Пошук у 2 етапи ------------------------
def match_code_two_stage(nom_text, allowed_category):
    """
    Етап 1: шукаємо тільки в allowed_category.
    Етап 2 (fallback): якщо не знайшли — шукаємо по інших субкатегоріях.

    Додаткова логіка:
    - Якщо 'Назва_для_аналізу' (довідник) складається лише з чисел/розділювачів,
      то підтверджуємо збіг ТІЛЬКИ коли ці числа в Номенклатурі:
        * НЕ прилягають до інших цифр зліва/справа (границі збігу),
        * і між послідовними числами є НЕцифровий символ.
    """
    search = normalize_and_replace(nom_text)
    raw_nom = '' if pd.isna(nom_text) else str(nom_text)

    # Етап 1 — всередині категорії
    if pd.notna(allowed_category):
        sub = cat_to_ref.get(str(allowed_category))
        if sub is not None and not sub.empty:
            for _, row in sub.iterrows():
                if not row['is_digits_only']:
                    if row['norm_name'] and row['norm_name'] in search:
                        return row['Внутрішній код']
                else:
                    # для числових назв — суворий збіг із розділювачами/межами
                    if digits_only_sequence_matches(raw_nom, row['digit_tokens']):
                        return row['Внутрішній код']

    # Етап 2 — по інших категоріях
    for _, row in df_ref_all.iterrows():
        if pd.notna(allowed_category) and str(row['Субкатегорія']) == str(allowed_category):
            continue
        if not row['is_digits_only']:
            if row['norm_name'] and row['norm_name'] in search:
                return row['Внутрішній код']
        else:
            if digits_only_sequence_matches(raw_nom, row['digit_tokens']):
                return row['Внутрішній код']

    return None

# ------------------------ 5.9) Маска "це запчастина" за колонкою Номенклатура ------------------------
mask_spare = df_nom['Номенклатура'].astype(str).str.contains(pattern_strict, regex=True, na=False)

# ------------------------ 6) Застосування до номенклатури ------------------------
codes = [
    match_code_two_stage(nm, cat)
    for nm, cat in zip(df_nom['Номенклатура'], df_nom['Культура'])
]

# Якщо це запчастина (mask_spare == True), "Внутрішній код" не повертаємо
df_nom['Внутрішній код'] = np.where(mask_spare, None, codes)

# ------------------------ 6.1) Додаємо «Загальна назва» з довідника ------------------------
# Беремо оригінальну «Загальна назва» із довідника (після фільтрів стоп-слів),
# але виключаємо коди, які ми вилучали у codes_to_remove
df_ref_lookup = (
    df_ref_raw.loc[~df_ref_raw['Внутрішній код'].isin(codes_to_remove), ['Внутрішній код', 'Загальна назва']]
    .drop_duplicates(subset=['Внутрішній код'])
)

# Ліве злиття до результату за кодом
df_nom = df_nom.merge(df_ref_lookup, on='Внутрішній код', how='left')

# ------------------------ 7) Фінальне впорядкування та збереження ------------------------
# Піднімемо службові та ключові колонки
cols = df_nom.columns.tolist()

# Якщо є колона з Номенклатурою+од.вим., ставимо її першою
leading = []
if 'Номенклатура з одиницею виміру' in cols:
    leading.append('Номенклатура з одиницею виміру')

# Приберемо службову колонку
if '_code_norm' in df_nom.columns:
    df_nom = df_nom.drop(columns=['_code_norm'])

# Запис Excel
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df_nom.to_excel(writer, sheet_name='номенклатура', index=False)

# ------------------------ 7.1) Підфарбувати колонки у синій ------------------------
from openpyxl import load_workbook
from openpyxl.styles import Font

wb = load_workbook(output_path)
ws = wb['номенклатура']

# Мапа заголовок -> індекс колонки (1-based)
header_to_idx = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

blue_font = Font(color="0000FF")  # синій колір

for col_name in ['Внутрішній код', 'Загальна назва']:
    col_idx = header_to_idx.get(col_name)
    if col_idx:
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_idx).font = "#A89064"

wb.save(output_path)

print('Готово! Результат збережено в', output_path)
